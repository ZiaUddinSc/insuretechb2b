from datetime import datetime
import re
from rest_framework.response import Response
from django.db.models import Count, Q, OuterRef, Subquery
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import permission_classes,authentication_classes,api_view
from django.contrib import messages
from django.http import HttpResponse
from b2bmanagement.models import Designation,Department,Organization,Plan,Bank,OrganizationPolicy
from accounts.models import CustomUser
from claim.models import EmployeeInformation,FileTransferHistory,ClaimInformation
from openpyxl.workbook import Workbook
from core.utils import get_last_six_or_pad, generate_auto_id,get_martial_status,get_relation_status,get_sex_status,get_bank_account_type_status,get_yes_no_status
from openpyxl import load_workbook
from rest_framework.permissions import IsAuthenticated
from django.http import JsonResponse
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import  Group
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework.authentication import TokenAuthentication
from django.db import transaction, IntegrityError
from django.db.models import (
    Max, Q,Sum, Count, F, DecimalField,
    ExpressionWrapper, Subquery, OuterRef
)
import openpyxl
# Create your views here.



@login_required(login_url='login')
def dashboard_counts(request,group_name="Super Admin"):
    employee = EmployeeInformation.objects.filter(user=request.user)
    try:
        employee = EmployeeInformation.objects.get(user=request.user)
    except EmployeeInformation.DoesNotExist:
        return {
            "organization": None,
            "status_summary": {},
            "show_system_gen_popup": "N"
        }

    current_contract_no = employee.contract_no
    
    # -------------------------------
    # Organization
    # -------------------------------
    organization = None
    policy = OrganizationPolicy.objects.filter(
        organization_contract_no=current_contract_no
    ).first()
    if policy:
        organization = Organization.objects.filter(
            id=policy.organization_id
        ).first()
    # -------------------------------
    # Base queryset
    # -------------------------------
    # Step 1: Base queryset
    if group_name=="B2B Employee":
        history_qs = FileTransferHistory.objects.filter(
            Q(sender=request.user)
        ).annotate(
            converted_amount=ExpressionWrapper(
                F('file__total_amount')  * F('file__exchange_rate'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
        
        
    elif group_name=="Organization HR":
        history_qs = FileTransferHistory.objects.filter(
            sender=request.user
            # file__employee__contract_no=current_contract_no
        ).annotate(
            converted_amount=ExpressionWrapper(
                F('file__total_amount')  * F('file__exchange_rate'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
    else:    
        history_qs = FileTransferHistory.objects.filter(
            from_group__name__iexact=group_name
            # ,file__employee__contract_no=current_contract_no
        ).annotate(
            converted_amount=ExpressionWrapper(
                F('file__total_amount') * F('file__exchange_rate'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
    status_summary = history_qs.values('status_after').annotate(
        count=Count('id'),
        total_amount=Sum('converted_amount')
    )
    # Step 3: Build lookup map (normalize keys)
    summary_map = {
        str(row['status_after']).lower(): {
            "count": row['count'],
            "total_amount": row['total_amount'] or 0
        }
        for row in status_summary
    }
    
    # Step 4: Build all statuses
    status_data = {}
    for status_code, status_label in ClaimInformation.FILE_TYPE_CHOICES:
        key = str(status_code)
        normalized_label = (
            status_label
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
        )

        status_data[normalized_label] = {
            "status_code": status_code,
            "count": summary_map.get(key, {}).get("count", 0),
            "total_amount": summary_map.get(key, {}).get("total_amount", 0)
        }
    selttled_status = ClaimInformation.objects.filter(
            created_by=request.user,file_status=3
     )    
    data = {
        "organization": organization.organization_name if organization else None,
        "status_summary": status_data,
        'show_system_gen_popup':request.user.system_generate,
        "selttled_count":selttled_status.count()
    }
    return data



@login_required(login_url='login')
def dashboard(request):
    dashboard_data=dashboard_counts(request,"B2B Employee")
    context = dashboard_data
    template_name = 'dashboard/dashboard.html'
    return render(request,template_name=template_name,context=context)

#Page Not Found View

def error_404_view(request, exception):
    return render(request, 'error-404.html')


def hr_admin_dashboard(request):
    dashboard_data=dashboard_counts(request,"Organization HR")
    context = dashboard_data
    return render(request, 'dashboard/hr_admin_dashboard.html',context=context)

def insurer_dashboard(request):
    dashboard_data=dashboard_counts(request,"Waada")
    context = dashboard_data
    return render(request, 'dashboard/insurer_dashboard.html',context=context)

def b2b_employee_dashboard(request):
    dashboard_data=dashboard_counts(request,"B2B Employee")
    context = dashboard_data
    return render(request, 'dashboard/b2b_employee_dashboard.html',context=context)

def insurer_claim_dashboard(request):
    dashboard_data=dashboardForClaimOfficerCount(request,"Insurer Claim Officer")
    context = dashboard_data
    return render(request, 'dashboard/insurer_claim_dashboard.html',context=context)

def insurer_audit_dashboard(request):
    dashboard_data=dashboardForClaimOfficerCount(request,"Insurer Audit Officer")
    context = dashboard_data
    return render(request, 'dashboard/insurer_audit_dashboard.html',context=context)

def insurer_supervisor_dashboard(request):
    dashboard_data=dashboardForClaimOfficerCount(request,"Claim Supervisor")
    context = dashboard_data
    return render(request, 'dashboard/insurer_supervisor_dashboard.html',context=context)

def insurer_finance_dashboard(request):
    dashboard_data=dashboard_counts(request,"Insurer Finance")
    context = dashboard_data
    return render(request, 'dashboard/insurer_finance_dashboard.html',context=context)

def my_policy(request):
    return render(request, 'dashboard/my_policy.html')


def waadaaSupervisor(request):
    return render(request, 'dashboard/waadaa_supervisor.html')


def required_document(request):
    return render(request, 'dashboard/required-document.html')

def dashboardCount(request,group_name="Organization HR"):
    latest_status = FileTransferHistory.objects.filter(
            to_group__name=group_name
     ).values('file').annotate(status=Max('status_after')) 
    latest_approved = FileTransferHistory.objects.filter(
            sender=request.user,from_group__name=group_name
     ).values('file').annotate(status_after=Max('status_after'))     
    # status_counts = latest_status.values('status').annotate(count=Count('file'))
    #DOC REQ
    latest_doc_req = FileTransferHistory.objects.filter(
            sender=request.user
     ).values('file').annotate(status=Max('status_after'))     # status_counts = latest_status.values('status').annotate(count=Count('file'))
    # FILE_TYPE_CHOICES = [
    #     (0, 'DECLINED'),
    #     (1, 'APPROVED'),
    #     (2, 'SUBMITTED'),
    #     (3, 'SETTELED'),
    #     (4, 'PENDING'),
    #     (5, 'PROCESSING'),
    #     (6, 'RETURN'),
    #     (7, 'DOCUMENT-WANTED')
    # ]
    total_files=latest_status.count()
    pending_count = latest_status.filter(status_after=4).count() #4=PENDING
    decline_count = latest_doc_req.filter(status_after=0).count() #4=DECLINEd
    approved_count = latest_approved.filter(status_after=4).count() #1=APPROVED
    doc_rec_count = latest_doc_req.filter(status_after=7).count() #4=DOCUMENT WANTED
    data = {
        'organization': '',
        'total_files': total_files,
        'approved_count': approved_count,
        'decline_count': decline_count,
        'pending_count': pending_count,
        'doc_rec_count':doc_rec_count,
        'show_system_gen_popup':request.user.system_generate
    }
    return data



def dashboardForClaimOfficerCount(request,group_name="Organization HR"):
    if group_name == "Claim Supervisor": 
        latest_status = FileTransferHistory.objects.filter(
                to_group__name=group_name
        ).values('file').annotate(status=Max('status_after')) 
    else:    
        latest_status = FileTransferHistory.objects.filter(
                to_group__name=group_name
        ).values('file').annotate(status=Max('status_after')) 
    
    stttled_status = FileTransferHistory.objects.filter(
                from_group__name=group_name
        ).values('file').annotate(status=Max('status_after')) 
    latest_approved = FileTransferHistory.objects.filter(
            sender=request.user,from_group__name=group_name
     ).values('file').annotate(status_after=Max('status_after'))     
    # status_counts = latest_status.values('status').annotate(count=Count('file'))
    #DOC REQ
    latest_doc_req = FileTransferHistory.objects.filter(
            sender=request.user
     ).values('file').annotate(status=Max('status_after'))     # status_counts = latest_status.values('status').annotate(count=Count('file'))
    # FILE_TYPE_CHOICES = [
    #     (0, 'DECLINED'),
    #     (1, 'APPROVED'),
    #     (2, 'SUBMITTED'),
    #     (3, 'SETTELED'),
    #     (4, 'PENDING'),
    #     (5, 'PROCESSING'),
    #     (6, 'RETURN'),
    #     (7, 'DOCUMENT-WANTED')
    # ]
    total_files=latest_status.count()
    pending_count = latest_status.filter(status_after=4).count() #4=PENDING
    decline_count = latest_doc_req.filter(status_after=0).count() #4=DECLINEd
    approved_count = latest_approved.filter(status_after=4).count() #1=APPROVED
    doc_rec_count = latest_doc_req.filter(status_after=7).count() #4=DOCUMENT WANTED
    settled_count = stttled_status.filter(status_after=3).count() #3=SETTLED
    data = {
        'organization': '',
        'total_files': total_files,
        'approved_count': approved_count,
        'decline_count': decline_count,
        'pending_count': pending_count,
        'doc_rec_count':doc_rec_count,
        "settled_count":settled_count,
        'show_system_gen_popup':request.user.system_generate
    }
    return data



def dashboardCountForOrganization(request,group_name="Organization HR"):
    
    try:
        employee = EmployeeInformation.objects.get(user=request.user)
    except EmployeeInformation.DoesNotExist:
        return {
        'organization':"",
        'total_files': 0,
        'approved_count': 0,
        'decline_count': 0,
        'pending_count': 0,
        'doc_rec_count':0,
        'show_system_gen_popup':""
    }
    organization=''
    if employee:
        current_contract_no = employee.contract_no
        employees_qs=EmployeeInformation.objects.filter(contract_no=employee.contract_no)
        curerent_policy=OrganizationPolicy.objects.filter(organization_contract_no=current_contract_no).first()
        if curerent_policy:
            organization=Organization.objects.filter(id=curerent_policy.organization_id).first()
    employee_user_list = employees_qs.exclude(user__id__isnull=True).values_list('user__id', flat=True)
    latest_status = FileTransferHistory.objects.filter(
            to_group__name=group_name
     ).values('file').annotate(status=Max('status_after')) 
    latest_approved = FileTransferHistory.objects.filter(
            receiver=request.user,to_group__name=group_name
     ).values('file').annotate(status_after=Max('status_after'))     
    # status_counts = latest_status.values('status').annotate(count=Count('file'))
    #DOC REQ
    latest_doc_req = FileTransferHistory.objects.filter(
            sender=request.user
     ).values('file').annotate(status=Max('status_after'))     # status_counts = latest_status.values('status').annotate(count=Count('file'))
    # FILE_TYPE_CHOICES = [
    #     (0, 'DECLINED'),
    #     (1, 'APPROVED'),
    #     (2, 'SUBMITTED'),
    #     (3, 'SETTELED'),
    #     (4, 'PENDING'),
    #     (5, 'PROCESSING'),
    #     (6, 'RETURN'),
    #     (7, 'DOCUMENT-WANTED')
    # ]
    total_files=latest_status.count()
    pending_count = latest_status.filter(status_after=4).count() #4=PENDING
    decline_count = latest_doc_req.filter(status_after=0).count() #4=DECLINEd
    approved_count = latest_approved.filter(status_after=1).count() #1=APPROVED
    doc_rec_count = latest_doc_req.filter(status_after=7).count() #4=DOCUMENT WANTED
    data = {
        'organization': organization.organization_name,
        'total_files': total_files,
        'approved_count': approved_count,
        'decline_count': decline_count,
        'pending_count': pending_count,
        'doc_rec_count':doc_rec_count,
        'show_system_gen_popup':request.user.system_generate
    }
    return data

@api_view(['GET'])
@authentication_classes([TokenAuthentication]) 
@permission_classes([IsAuthenticated])
def dashboardCountForCustomer(request):
    GROUP_NAME = "B2B Employee"

    # -------------------------------
    # Employee
    # -------------------------------
    try:
        employee = EmployeeInformation.objects.get(user=request.user)
    except EmployeeInformation.DoesNotExist:
        return Response({"success": False, "data": {}})

    current_contract_no = employee.contract_no

    # -------------------------------
    # Organization
    # -------------------------------
    organization = None
    policy = OrganizationPolicy.objects.filter(
        organization_contract_no=current_contract_no
    ).first()

    if policy:
        organization = Organization.objects.filter(
            id=policy.organization_id
        ).first()

    # -------------------------------
    # Latest status subquery
    # -------------------------------
    latest_status_subquery = (
        FileTransferHistory.objects
        .filter(
            file_id=OuterRef('pk'),
            to_group__name=GROUP_NAME
        )
        .order_by('-id')   # or '-created_at'
        .values('status_after')[:1]
    )

    # -------------------------------
    # Base queryset
    # -------------------------------
    claims_qs = (
        ClaimInformation.objects
        .filter(employee__contract_no=current_contract_no)
        .annotate(
            latest_status=Subquery(latest_status_subquery),
            converted_amount=ExpressionWrapper(
                F('total_amount') * F('exchange_rate'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
    )

    # -------------------------------
    # STATUS-WISE COUNT + AMOUNT
    # -------------------------------
    status_summary = (
        claims_qs
        .values('latest_status')
        .annotate(
            count=Count('id'),
            total_amount=Sum('converted_amount')
        )
    )

    STATUS_MAP = dict(ClaimInformation.FILE_TYPE_CHOICES)

    status_data = {
        STATUS_MAP.get(row['latest_status'], 'UNKNOWN').lower(): {
            "status_code": row['latest_status'],
            "count": row['count'],
            "total_amount": row['total_amount'] or 0
        }
        for row in status_summary
    }

    # -------------------------------
    # Response
    # -------------------------------
    return Response({
        "success": True,
        "data": {
            "organization": organization.organization_name if organization else None,
            "status_summary": status_data,
            "show_system_gen_popup": request.user.system_generate
        }
    }) 
def GroupClaim(request):
    dashboard_data=dashboardCountForOrganization(request)
    context = dashboard_data
    template_name = 'b2bproduct/group-claim.html'
    return render(request, template_name=template_name,context=context)


def WaadaaOperationDashboard(request):
    dashboard_data=dashboardCount(request,"Shield Operation")
    context = dashboard_data
    template_name = 'dashboard/waadaa_operation_dashboard.html'
    return render(request, template_name=template_name,context=context)


@login_required
def download_excel_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Information"

    headers = ['Org_name', 'Employee_Name', 
               'Employee_ID', 'Member_Name',
               'Relation','Nominee',
               'Plan','Designation',
               'Department','Emp_Email','DOB',
               'Age','Sex',
               'Maritial_Status','Sum_Assured_Life',
               'A/C_Type','Bank_Name','Ac_NO','Routing_No',
               'Mat_Status','Mat_Plan',
              ]
    ws.append(headers)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employee-template.xlsx"'

    wb.save(response)
    return response

# @transaction.atomic
def process_data(request,rows):
    try:
        with transaction.atomic():
            new_org_auto_id = generate_auto_id(Organization,3)
            contract_no=f"{new_org_auto_id}{'-A'}"
            designation_names =set()
            department_names= set()
            plan_names= set()
            organization_names=set()
            bank_names=set()
            inserted = []
            duplicates = []
            temp_emp=''
            empr_ids=[]
            users_data = []       
            useremails = set()   
            group_name = "B2B Employee"  # or get from request.POST
            hr_group_name = "Organization HR"  # or get from request.POST
            group, created = Group.objects.get_or_create(name=group_name)
            hr_group, created = Group.objects.get_or_create(name=hr_group_name)
            child_counter = {}
            current_id = None
            relation_code = {
                "Self": 0,
                "Father": 8,
                "Mother": 9,
                "Wife": 1,   # If you confirm
            }
            # rows = sorted(rows, key=lambda x: x["employee_id"])
            for row in rows:
                id_ = str(row["employee_id"]).strip()
                relation = row["relation"].strip()
                code=''
                # Reset child counter only when NEW employee_id
                if id_ != current_id:
                    current_id = id_
                    child_counter[id_] = 2     # child ALWAYS starts from 2

                # Child logic
                if relation == "Child":
                    code = child_counter[id_]
                    child_counter[id_] += 1
                else:
                    code = relation_code.get(relation)
                emp_id = f"{id_}-{code}"
                if(row['designation']):
                    designation_names.add(row['designation'].strip())
                if(row['department']):
                    department_names.add(row['department'].strip())
                if(row['plan']):
                    plan_names.add(row['plan'].strip())
                if(row['org_name']):
                    organization_names.add(row['org_name'].strip())
                if(row['bank_name']):
                    bank_names.add(row['bank_name'].strip())
                if row['employee_name']:
                    temp_emp=row['employee_id']
                if emp_id:    
                    empr_ids.append(emp_id)
                new_member_id=f"{new_org_auto_id}{'00'}{emp_id}"
                if  str(row['emp_email']).strip():
                    username = str(row['member_name']).strip()
                    email = str(row['emp_email']).strip() 
                    mobile_no = str(row['mobile_no']).strip() 
                    hr_admin = str(row['hr_admin']).strip() 
                    
                    if str(hr_admin).strip() in ['Yes','y','YES']:
                        groups=[group,hr_group]    
                    else:
                        groups=[group]    
                if email in useremails or email in [None, "", "None"]:  # skip duplicates in Excel
                    continue
                useremails.add(email)
                users_data.append({
                    "username": username,
                    "email": email,
                    "phone_number": mobile_no,
                    'groups': groups
                })     
            #Filter out users already in DB
            existing_emails = set(CustomUser.objects.filter(
                email__in=[u['email'] for u in users_data]
            ).values_list('email', flat=True))
            new_users = [
                    CustomUser(
                        username=u['username'],
                        email=u['email'],
                        system_generate='Y',
                        password=make_password('123456'),  # hash password
                        is_staff=True,
                        is_active=True,
                        is_status=True
                        )
                    for u in users_data if u['email'] not in existing_emails
                ]
            CustomUser.objects.bulk_create(new_users)
            created_users = CustomUser.objects.filter(email__in=useremails)
            user_map = {u.email: u for u in created_users}

            #  Map all users
            all_users = {u.email: u for u in CustomUser.objects.filter(
                email__in=[ud['email'] for ud in users_data]
            )}
            for u in users_data:
                user_obj = user_map[u['email']]
                groups_qs= u['groups']   # This is a list like ['staff','manager']
                user_obj.groups.set(groups_qs)
                
            designation_map = {c.title: c for c in Designation.objects.filter(title__in=designation_names)}
            department_map = {s.name: s for s in Department.objects.filter(name__in=department_names)}
            plan_map = {s.name: s for s in Plan.objects.filter(name__in=plan_names)}
            # organization_map = {s.organization_name: s 
            #                     for s in Organization.objects.filter(organization_name__in=organization_names)
            #                     }
            bank_map = {s.name: s for s in Bank.objects.filter( Q(name__in=bank_names) |  Q(short_name__in=bank_names))}
            # ====== 3. Create missing FKs ======
            new_designations = [Designation(title=title) for title in designation_names if title not in designation_map]
            new_departments = [Department(name=name) for name in department_names if name not in department_map]
            new_plans = [Plan(name=name) for name in plan_names if name not in plan_map]
            # new_organizations = [Organization(organization_name=organization_name,organization_code=new_org_auto_id) for organization_name in organization_names if organization_name not in organization_map]
            new_banks = [Bank(name=name) for name in bank_names if name not in bank_map]
            Designation.objects.bulk_create(new_designations)
            Department.objects.bulk_create(new_departments)
            Plan.objects.bulk_create(new_plans)
            # Organization.objects.bulk_create(new_organizations)
            Bank.objects.bulk_create(new_banks)
            # Refresh FK maps after insert
            designation_map = {c.title: c for c in Designation.objects.filter(title__in=designation_names)}
            department_map = {s.name: s for s in Department.objects.filter(name__in=department_names)}
            plan_map = {c.name: c for c in Plan.objects.filter(name__in=plan_names)}
            # organization_map = {s.organization_name: s for s in Organization.objects.filter(organization_name__in=organization_names)}
            bank_map = {s.name: s for s in Bank.objects.filter(name__in=bank_names)}
            existing_employees = EmployeeInformation.objects.filter(
               (Q(employee_id__in=empr_ids) & Q(organization_emp_policy__organization_id=request.POST.get('organization'))) | Q(employee_email__in=useremails)
            ).values('employee_id', 'employee_email')
            existing_ids = {e['employee_id'] for e in existing_employees}
            existing_email_employee = {e['employee_email'] for e in existing_employees}
            employees_to_create = []
           
            for row in rows:   
                if str(row['employee_id']).strip():
                    temp_emp_id=get_last_six_or_pad(str(row['employee_id']).strip())
                id_ = str(row["employee_id"]).strip()
                relation = row["relation"].strip()
                code=''
                # Reset child counter only when NEW employee_id
                if id_ != current_id:
                    current_id = id_
                    child_counter[id_] = 2     # child ALWAYS starts from 2

                # Child logic
                if relation == "Child":
                    code = child_counter[id_]
                    child_counter[id_] += 1
                else:
                    code = relation_code.get(relation)
                emp_id = f"{id_}-{code}"    
                
                new_member_id=f"{new_org_auto_id}{'00'}{emp_id}"
                
                if row['dob']:
                    dob = datetime.strptime(row['dob'], "%d-%b-%y").date()  
                if row['membership_date']:
                    membership_date = datetime.strptime(row['membership_date'], "%d-%b-%y").date()   
                row_plan=row['plan'].strip()
                row_designation=row['designation'].strip()
                row_department=row['department'].strip()
                row_bank=row['bank_name'].strip()
                row_org_name=row['org_name'].strip()
                department = department_map.get(row_department)
                designation = designation_map.get(row_designation)
                plan = plan_map.get(row_plan)
                bank = bank_map.get(row_bank)
                # organization = organization_map.get(row_org_name)
                sex_status= get_sex_status(row['sex'].strip())
                martial_status= get_martial_status(row['maritial_status'].strip())
                ac_status= get_bank_account_type_status(row['a/c_type'].strip())
                relation=get_relation_status(str(row['relation']).strip())
                hr_admin=get_yes_no_status(str(row['hr_admin']).strip())
                salary=0
                sum_assured=0
                nominee_type='n'
                if row['nominee']=='Yes':
                    nominee_type='y'
                if row['salary']:
                    salary=row['salary']
                if row['sum_assured_life']:
                    sum_assured=row['sum_assured_life']
                email = str(row['emp_email']).strip() 
                if emp_id in existing_ids or (email and email in existing_email_employee):
                    duplicates.append({
                        "employee_id": emp_id,
                        "member_name": row['member_name'],
                        "email": row['emp_email'],
                        "status": "Duplicate"
                    })
                else:
                    user_obj=None
                    email = str(row['emp_email']).strip() 
                    if email:
                        user_obj = all_users.get(email)
                    org_id =request.POST.get('organization')
                    org_contract_id=request.POST.get('org_contract_id')
                    org_policy = OrganizationPolicy.objects.filter(organization__id=org_id,organization_contract_no=org_contract_id).first()
                    if org_policy:
                        org_id=org_policy.id
                    else:
                        org_policy = OrganizationPolicy.objects.create(organization_id=org_id,organization_contract_no=org_contract_id)
                        org_id=org_policy.id
                    employees_to_create.append(
                        EmployeeInformation(
                            organization_emp_policy_id=org_id,
                            contract_no=org_contract_id,
                            member_id=new_member_id,
                            designation=designation,
                            user=user_obj,
                            department=department,
                            member_name=str(row['member_name']).strip(),
                            employee_name=str(row['employee_name']).strip(),
                            bank=bank,
                            mobile_number=str(row['mobile_no']).strip(),
                            employee_email=str(row['emp_email']).strip(),
                            age=row['age'], 
                            dob=dob,
                            nominee_type=nominee_type,
                            salary=salary,
                            membership_date=membership_date,
                            plan=plan,
                            employee_id=emp_id,
                            sum_assured_life=sum_assured,
                            branch_name=str(row['branch_name']).strip(),
                            marital_status=martial_status,
                            sex=sex_status,
                            account_number=str(row['ac_no']).strip(),
                            relation=relation,
                            ac_type=ac_status,
                            hr_type=hr_admin,
                            created_by=request.user
                        )
                    )
                inserted.append({
                    "member_id": new_member_id,
                    "member_name": row['member_name'],
                    "status": "Inserted"
                })
            if employees_to_create:
                EmployeeInformation.objects.bulk_create(employees_to_create)
            return {"status": True,"inserted": inserted, "duplicates": duplicates}
    except IntegrityError as e:
        return {'status': False, 'message': 'Database error: ' + str(e)}
    except Exception as e:
        return {'status': False, 'message': 'Unexpected error: ' + str(e)}


@login_required
def upload_excel(request):
    data = []
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']    
        data = []  
        try:         
            wb = load_workbook(excel_file, data_only=False)
        except InvalidFileException:
            return JsonResponse({"error": "Invalid Excel file format. Please upload a valid .xlsx file."}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error reading Excel file: {str(e)}"}, status=500)
        sheet = wb.active  # or wb[wb.sheetnames[0]]
        if "Sheet1" in wb.sheetnames:
            sheet = wb["Sheet1"]
        else:
            # fallback: pick the first sheet if Sheet1 doesn't exist
            sheet = wb[wb.sheetnames[0]]

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            # Empty sheet
            data = []
        else:
            headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
            for row in rows[1:]:
                row_dict = {header: value for header, value in zip(headers, row)}
                data.append(row_dict)
            result=process_data(request,data)  
            return JsonResponse(result)
    return JsonResponse({"error": "No file uploaded"}, status=400)